from fastapi import APIRouter, Depends, HTTPException, status, Request, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from uuid import UUID
from typing import Optional, List, Union
from datetime import date, datetime
import io

from app.db.base import get_db
from app.services.survey_service import SurveyService
from app.schemas.survey import SurveyResponse, SurveyCreate
from app.schemas.response import SurveyResponseCreate, SurveyResponseResponse
from app.api.dependencies import get_current_user, get_current_admin, get_current_account, get_current_regular_user
from app.services import membership_service
from app.models.user import User
from app.models.admin import Admin
from app.models.client import Client
from pydantic import BaseModel

router = APIRouter()


class ToggleSurveyRequest(BaseModel):
    is_active: bool


def ensure_can_view_survey_results(survey, account: Union[User, Admin, Client]) -> None:
    """Verifica que la cuenta pueda ver resultados/analytics de la encuesta.

    Reglas: los admin ven todo; los clientes solo sus propias encuestas;
    los ciudadanos (User) no tienen acceso a analytics.
    Lanza 403 si no está autorizado.
    """
    if isinstance(account, Admin):
        return
    if isinstance(account, Client) and survey.client_id == account.id:
        return
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="No tenés permisos para ver los datos de esta encuesta"
    )


@router.get("/", response_model=List[SurveyResponse])
def get_surveys(
    current_user: Union[User, Admin, Client] = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Obtiene las encuestas según el tipo de cuenta:
    - Admin: ve todas las encuestas
    - Client: ve solo sus encuestas
    - User: no tiene acceso a esta ruta
    """
    if isinstance(current_user, Admin):
        # Admin ve todas las encuestas
        surveys = SurveyService.get_all_surveys(db)
    elif isinstance(current_user, Client):
        # Cliente ve solo sus encuestas
        surveys = SurveyService.get_all_surveys(db, client_id=current_user.id)
    else:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para acceder a esta funcionalidad"
        )

    return surveys


@router.get("/active", response_model=SurveyResponse)
def get_active_survey(db: Session = Depends(get_db)):
    """Obtiene la encuesta activa actual (DEPRECADO: global, sin scope por cliente).

    Usar GET /surveys/available para el flujo del ciudadano.
    """
    survey = SurveyService.get_active_survey(db)
    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No hay encuesta activa disponible"
        )
    return survey


@router.get("/available", response_model=List[SurveyResponse])
def get_available_surveys(
    current_user: User = Depends(get_current_regular_user),
    db: Session = Depends(get_db),
):
    """Encuestas que el ciudadano logueado puede responder.

    Incluye las públicas y las de los municipios a los que pertenece (con herencia
    por jerarquía de clientes). Reemplaza el uso ciudadano de /active.
    """
    return SurveyService.get_available_surveys(db, current_user)


@router.get("/participation-trend")
def get_participation_trend(
    gender: Optional[str] = None,
    survey_id: Optional[str] = None,
    age_range: Optional[str] = None,
    current_user: Union[User, Admin, Client] = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Devuelve la tendencia de participación mensual (respuestas por mes).
    Si se pasa survey_id, filtra solo esa survey. Si no, muestra todas las del cliente.
    Opcionalmente filtrar por género (masculino/femenino) y/o rango de edad (18-30, 31-45, 46-60, 60+).
    """
    from app.models.response import SurveyResponse as SurveyResponseModel
    from app.models.survey import Survey

    if survey_id:
        # Verificar que la cuenta pueda ver esta encuesta puntual (evita fuga cross-tenant)
        survey = SurveyService.get_survey_by_id(db, UUID(survey_id))
        if not survey:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Encuesta no encontrada")
        ensure_can_view_survey_results(survey, current_user)
        target_survey_ids = [survey_id]
    else:
        if isinstance(current_user, Client):
            target_survey_ids = [str(s.id) for s in db.query(Survey.id).filter(Survey.client_id == current_user.id).all()]
        elif isinstance(current_user, Admin):
            target_survey_ids = [str(s.id) for s in db.query(Survey.id).all()]
        else:
            raise HTTPException(status_code=403, detail="No tienes permisos")

    if not target_survey_ids:
        return {"months": []}

    query = (
        db.query(
            extract("year", SurveyResponseModel.completed_at).label("year"),
            extract("month", SurveyResponseModel.completed_at).label("month"),
            func.count(SurveyResponseModel.id).label("count"),
        )
        .filter(
            SurveyResponseModel.survey_id.in_(target_survey_ids),
            SurveyResponseModel.completed == True,
            SurveyResponseModel.completed_at.isnot(None),
        )
    )

    needs_user_join = False

    if gender and gender.lower() in ("masculino", "femenino"):
        needs_user_join = True

    if age_range and age_range in ("18-30", "31-45", "46-60", "60+"):
        needs_user_join = True

    if needs_user_join:
        query = query.join(User, User.id == SurveyResponseModel.user_id)

        if gender and gender.lower() in ("masculino", "femenino"):
            query = query.filter(func.lower(User.gender) == gender.lower())

        if age_range:
            age_expr = extract("year", func.age(func.now(), User.birth_date))
            if age_range == "18-30":
                query = query.filter(User.birth_date.isnot(None), age_expr >= 18, age_expr <= 30)
            elif age_range == "31-45":
                query = query.filter(User.birth_date.isnot(None), age_expr >= 31, age_expr <= 45)
            elif age_range == "46-60":
                query = query.filter(User.birth_date.isnot(None), age_expr >= 46, age_expr <= 60)
            elif age_range == "60+":
                query = query.filter(User.birth_date.isnot(None), age_expr > 60)

    results = query.group_by("year", "month").order_by("year", "month").all()

    # Construir lista de los últimos 12 meses con datos
    now = datetime.now()
    month_names = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    result_map = {(int(r.year), int(r.month)): int(r.count) for r in results}

    months = []
    for i in range(11, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        count = result_map.get((y, m), 0)
        months.append({
            "label": month_names[m - 1],
            "year": y,
            "month": m,
            "count": count,
        })

    # Calcular tendencia mensual (cambio % vs mes anterior)
    current_month_count = months[-1]["count"] if months else 0
    prev_month_count = months[-2]["count"] if len(months) >= 2 else 0
    if prev_month_count > 0:
        trend_pct = round(((current_month_count - prev_month_count) / prev_month_count) * 100)
    else:
        trend_pct = 100 if current_month_count > 0 else 0

    return {
        "months": months,
        "current_month": current_month_count,
        "previous_month": prev_month_count,
        "trend_percentage": trend_pct,
    }


@router.get("/{survey_id}", response_model=SurveyResponse)
def get_survey(survey_id: UUID, db: Session = Depends(get_db)):
    """Obtiene una encuesta por ID"""
    survey = SurveyService.get_survey_by_id(db, survey_id)
    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Encuesta no encontrada"
        )
    return survey


@router.post("/", response_model=SurveyResponse, status_code=status.HTTP_201_CREATED)
def create_survey(
    survey_data: SurveyCreate,
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Crea una nueva encuesta (solo Admin)"""
    try:
        survey = SurveyService.create_survey(db, survey_data)
        return survey
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@router.post("/responses", response_model=SurveyResponseResponse, status_code=status.HTTP_201_CREATED)
def submit_survey_response(
    response_data: SurveyResponseCreate,
    request: Request,
    current_user: User = Depends(get_current_regular_user),
    db: Session = Depends(get_db)
):
    """Envía una respuesta de encuesta (requiere ciudadano autenticado)."""
    try:
        # Verificar que la encuesta esté activa
        survey = SurveyService.get_survey_by_id(db, response_data.survey_id)
        if not survey:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Encuesta no encontrada"
            )
        if not survey.is_active:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Esta encuesta no está disponible actualmente"
            )

        # Elegibilidad: pública o miembro del municipio dueño
        if not membership_service.user_can_access_survey(db, current_user, survey):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="No estás habilitado para responder esta consulta"
            )

        # Capturar IP y User Agent
        ip_address = request.client.host if request.client else None
        user_agent = request.headers.get("user-agent")

        response = SurveyService.submit_response(
            db,
            response_data,
            user_id=current_user.id,
            ip_address=ip_address,
            user_agent=user_agent
        )
        return response
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar la respuesta: {str(e)}"
        )


@router.get("/can-respond/{survey_id}/{user_id}")
def check_can_respond(
    survey_id: UUID,
    user_id: UUID,
    current_account: Union[User, Admin, Client] = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Verifica si un usuario puede responder una encuesta (elegibilidad + límite).

    Solo el propio usuario o un admin pueden consultarlo (evita enumeración de ciudadanos).
    """
    if not (isinstance(current_account, Admin) or
            (isinstance(current_account, User) and current_account.id == user_id)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tenés permisos para consultar este recurso"
        )

    survey = SurveyService.get_survey_by_id(db, survey_id)
    user = db.query(User).filter(User.id == user_id).first()

    if not survey or not user:
        return {"can_respond": False, "message": "Encuesta o usuario no encontrado"}

    if not membership_service.user_can_access_survey(db, user, survey):
        return {"can_respond": False, "message": "No estás habilitado para responder esta consulta"}

    can_respond = SurveyService.user_can_respond(db, user_id, survey_id)
    return {
        "can_respond": can_respond,
        "message": "Usuario puede responder" if can_respond else "Ya alcanzaste el límite de respuestas para esta encuesta"
    }


@router.get("/{survey_id}/results")
def get_survey_results(
    survey_id: UUID,
    date_from: Optional[date] = Query(None, description="Fecha inicio del período (YYYY-MM-DD)"),
    date_to: Optional[date] = Query(None, description="Fecha fin del período (YYYY-MM-DD)"),
    current_user: Union[User, Admin, Client] = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Obtiene los resultados y estadísticas de una encuesta.
    Solo accesible para admin y cliente dueño de la encuesta.
    Opcionalmente filtra por período con date_from y date_to.
    """
    # Verificar que la encuesta existe
    survey = SurveyService.get_survey_by_id(db, survey_id)
    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Encuesta no encontrada"
        )

    # Verificar permisos
    if isinstance(current_user, User):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para ver resultados"
        )

    if isinstance(current_user, Client) and survey.client_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para ver esta encuesta"
        )

    # Obtener resultados
    results = SurveyService.get_survey_results(db, survey_id, date_from=date_from, date_to=date_to)
    return results


@router.patch("/{survey_id}/toggle")
def toggle_survey_status(
    survey_id: UUID,
    toggle_data: ToggleSurveyRequest,
    current_account: Union[Admin, Client] = Depends(get_current_account),
    db: Session = Depends(get_db)
):
    """
    Activa o desactiva una encuesta.
    Accesible para admin y client.
    """
    if not isinstance(current_account, (Admin, Client)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para esta acción"
        )

    # Verificar que la encuesta existe
    survey = SurveyService.get_survey_by_id(db, survey_id)
    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Encuesta no encontrada"
        )

    # Un cliente solo puede activar/desactivar sus propias encuestas (evita tocar otros municipios)
    if isinstance(current_account, Client) and survey.client_id != current_account.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tenés permisos para modificar esta encuesta"
        )

    # Actualizar estado
    survey.is_active = toggle_data.is_active
    survey.status = "active" if toggle_data.is_active else "inactive"
    db.commit()
    db.refresh(survey)

    return {
        "id": str(survey.id),
        "is_active": survey.is_active,
        "message": f"Encuesta {'activada' if toggle_data.is_active else 'desactivada'} exitosamente"
    }


@router.get("/{survey_id}/segments")
def get_survey_segments(
    survey_id: UUID,
    threshold: int = Query(20, ge=1, le=100, description="Umbral mínimo de % para incluir en segmento"),
    current_user: Union[User, Admin, Client] = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Obtiene la segmentación de votantes por preferencias.
    Agrupa usuarios que asignaron >= threshold% a cada área.
    """
    survey = SurveyService.get_survey_by_id(db, survey_id)
    if not survey:
        raise HTTPException(status_code=404, detail="Encuesta no encontrada")

    if isinstance(current_user, User):
        raise HTTPException(status_code=403, detail="No tienes permisos")

    if isinstance(current_user, Client) and survey.client_id != current_user.id:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver esta encuesta")

    return SurveyService.get_survey_segments(db, survey_id, threshold=threshold)


@router.get("/{survey_id}/segments/export")
def export_survey_segments(
    survey_id: UUID,
    threshold: int = Query(20, ge=1, le=100, description="Umbral mínimo de % para incluir en segmento"),
    current_user: Union[User, Admin, Client] = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Exporta los segmentos a un archivo XLSX con un tab por segmento.
    """
    survey = SurveyService.get_survey_by_id(db, survey_id)
    if not survey:
        raise HTTPException(status_code=404, detail="Encuesta no encontrada")

    if isinstance(current_user, User):
        raise HTTPException(status_code=403, detail="No tienes permisos")

    if isinstance(current_user, Client) and survey.client_id != current_user.id:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver esta encuesta")

    data = SurveyService.get_survey_segments(db, survey_id, threshold=threshold)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")

    try:
        wb = Workbook()
        # Eliminar la hoja por defecto
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        for segment in data["segments"]:
            # Nombre de hoja (max 31 chars para Excel)
            sheet_name = segment["area"][:31]
            ws = wb.create_sheet(title=sheet_name)

            # Headers
            headers = ["Nombre", "Email", "Barrio", "Ciudad", "% Asignado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Datos
            for row_idx, user in enumerate(segment["users"], 2):
                ws.cell(row=row_idx, column=1, value=user["name"])
                ws.cell(row=row_idx, column=2, value=user["email"])
                ws.cell(row=row_idx, column=3, value=user["neighborhood"])
                ws.cell(row=row_idx, column=4, value=user["city"])
                ws.cell(row=row_idx, column=5, value=f"{user['percentage']}%")

            # Auto-ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        # Si no hay segmentos, crear hoja vacía
        if not data["segments"]:
            ws = wb.create_sheet(title="Sin segmentos")
            ws.cell(row=1, column=1, value="No se encontraron segmentos con el umbral seleccionado")

        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        import unicodedata
        safe_title = unicodedata.normalize("NFKD", survey.title[:30]).encode("ascii", "ignore").decode("ascii")
        safe_title = safe_title.replace(" ", "_").replace("/", "_").replace("\\", "_")
        filename = f"segmentos_{safe_title}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar XLSX: {str(e)}")
